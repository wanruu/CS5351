from openpyxl import load_workbook

wb = load_workbook(filename=r'C:\Users\55429\Desktop\Data Engineering\HW1\Question3Code\information.xlsx')  ##读取路径
ws = wb.get_sheet_by_name("random_information")  ##读取名字为Sheet1的sheet表
num = 1

while 1:
    cell = ws.cell(row=num, column=1).value
    if cell:
        num = num + 1
    else:
        print(num)
        exit()